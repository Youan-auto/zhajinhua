import openpyxl
import datetime
import re
from unittest import TestCase

ts = TestCase()

file = openpyxl.load_workbook("D:\\Adobe\\美的项目\\新建文件夹\\测试报告11.28.xlsx")
bugs_list_sheet = file["Bugs"]
sit_summary_sheet = file["SIT summary"]
uat_summary_sheet = file["UAT summary"]
max_row = bugs_list_sheet.max_row
today = str(datetime.date.today())

bugs_list = []
for i in range(2, max_row + 1):
    one_row_list = []
    for j in range(1, 8):
        one_row_data = bugs_list_sheet.cell(row=i, column=j).value
        one_row_list.append(one_row_data)
    bugs_list.append(one_row_list)

day_sit_count = 1
day_uat_count = 1
sit_day = None
uat_day = None
sit_create_count = None
uat_create_count = None
close_count = 0
uat_count = 0
sit_closed_date_list = []
uat_closed_date_list = []
un_close_sit = 0
un_close_uat = 0
for bug in bugs_list:
    tfs_id = bug[1]
    des = bug[2]
    status = bug[3]
    env = bug[4]
    create_date = bug[5]
    create_ymd = re.search("2.+\\ ", str(create_date)).group()
    closed_date = bug[6]
    if env == "SIT" or not env or env == "Not Defect; SIT" or env == "Not Defect":
        if closed_date:
            closed_date = re.search("2.+\\ ", str(closed_date)).group()
            sit_closed_date_list.append(closed_date)
        if status != "Closed":
            un_close_sit += 1
        if sit_day != create_ymd:
            sit_create_count = 1
            day_sit_count += 1
            sit_day = create_ymd
            sit_summary_sheet.cell(row=day_sit_count, column=1, value=sit_day)
            sit_summary_sheet.cell(row=day_sit_count, column=2, value=sit_create_count)
        else:
            sit_create_count += 1
            sit_summary_sheet.cell(row=day_sit_count, column=2, value=sit_create_count)

    if env == "UAT" or env == "enhancement; UAT":
        if closed_date:
            closed_date = re.search("2.+\\ ", str(closed_date)).group()
            uat_closed_date_list.append(closed_date)
        if status != "Closed":
            un_close_uat += 1
        if uat_day != create_ymd:
            uat_create_count = 1
            day_uat_count += 1
            uat_day = create_ymd
            uat_summary_sheet.cell(row=day_uat_count, column=1, value=uat_day)
            uat_summary_sheet.cell(row=day_uat_count, column=2, value=uat_create_count)
        else:
            uat_create_count += 1
            uat_summary_sheet.cell(row=day_uat_count, column=2, value=uat_create_count)

file.save("D:\\Adobe\\美的项目\\新建文件夹\\测试报告11.28.xlsx")

sit_summary_list = []
sit_summary_sheet_max_row = sit_summary_sheet.max_row
for i in range(2, max_row + 1):
    one_row_list_01 = []
    for j in range(1, 3):
        one_row_data_01 = sit_summary_sheet.cell(row=i, column=j).value
        if one_row_data_01:
            one_row_list_01.append(one_row_data_01)
    if len(one_row_list_01) > 0:
        sit_summary_list.append(one_row_list_01)

sit_count = 1
for data in sit_summary_list:
    h = 0
    for closed_date in sit_closed_date_list:
        if closed_date == data[0]:
            h += 1
    sit_count += 1
    sit_summary_sheet.cell(row=sit_count, column=1, value=data[0])
    sit_summary_sheet.cell(row=sit_count, column=2, value=data[1])
    sit_summary_sheet.cell(row=sit_count, column=3, value=h)

sit_summary_sheet.cell(row=sit_count, column=4, value=un_close_sit)

file.save("D:\\Adobe\\美的项目\\新建文件夹\\测试报告11.28.xlsx")

uat_summary_list = []
uat_summary_sheet_max_row = sit_summary_sheet.max_row
for i in range(2, max_row + 1):
    one_row_list_01 = []
    for j in range(1, 3):
        one_row_data_01 = uat_summary_sheet.cell(row=i, column=j).value
        if one_row_data_01:
            one_row_list_01.append(one_row_data_01)
    if len(one_row_list_01) > 0:
        uat_summary_list.append(one_row_list_01)

uat_count = 1
print(uat_closed_date_list)
for data in uat_summary_list:
    h = 0
    for closed_date in uat_closed_date_list:
        if closed_date == data[0]:
            h += 1
    uat_count += 1
    uat_summary_sheet.cell(row=uat_count, column=1, value=data[0])
    uat_summary_sheet.cell(row=uat_count, column=2, value=data[1])
    uat_summary_sheet.cell(row=uat_count, column=3, value=h)

uat_summary_sheet.cell(row=uat_count, column=4, value=un_close_uat)


file.save("D:\\Adobe\\美的项目\\新建文件夹\\测试报告11.28.xlsx")
